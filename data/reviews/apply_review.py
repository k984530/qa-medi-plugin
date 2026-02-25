import openpyxl, json, sys, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load original
src_path = r'c:\Users\kjeon\QA\qa_agent\[Medisolve] centurion Say TestCase(상담관제) (2).xlsx'
wb = openpyxl.load_workbook(src_path)
ws_src = wb['Admin-AI가이드대시보드_자동화 초안']

# Load review JSON
with open(r'c:\Users\kjeon\QA\qa_agent\data\reviews\tc_review_v1.3.0_ai_guide_dashboard.json', 'r', encoding='utf-8') as f:
    review = json.load(f)

# Create new sheet (copy approach: read all data, apply fixes, write to new sheet)
ws = wb.create_sheet('Admin-AI가이드대시보드_리뷰반영')

# Styles
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='top')

def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.border = copy(src_cell.border)
    dst_cell.number_format = src_cell.number_format

# Step 1: Read all source rows into list of dicts
src_rows = []
for row in ws_src.iter_rows(min_row=1, max_row=202, max_col=15):
    row_data = {}
    row_styles = {}
    for cell in row:
        row_data[cell.column] = cell.value
        row_styles[cell.column] = cell
    src_rows.append({'data': row_data, 'styles': row_styles, 'row_num': row[0].row})

# Step 2: Apply fixes to source data
# V-010: Row 18 (index 17) - fix comma in Expected Result
for r in src_rows:
    if r['row_num'] == 18:
        r['data'][9] = '리스트 영역 접힘\n- 타이틀만 표시됨'
        break

# V-011: Row 50 (index 49) - split condition from Expected Result
for r in src_rows:
    if r['row_num'] == 50:
        r['data'][6] = '관리자 A가 알림 확인'  # F column (5Depth)
        r['data'][9] = '관리자 B에게도 해당 알림 즉시 제거됨'  # I column
        break

# V-001: Row 118 - mark for deletion (duplicate of 119)
delete_rows = {118}

# Step 3: Define new TCs to insert (after specific rows)
new_tcs = {
    16: [  # QA-001: after row 16
        {'B': '', 'C': '', 'D': '', 'E': '점검 필요 항목 4개 이하', 'F': '', 'G': '', 'H': 'P2', 'I': '해당 건수만큼 모두 표시됨', 'O': '리뷰 반영: QA-001 경계값'},
        {'B': '', 'C': '', 'D': '', 'E': '점검 필요 항목 5개', 'F': '', 'G': '', 'H': 'P2', 'I': '5개 모두 표시됨', 'O': '리뷰 반영: QA-001 경계값'},
    ],
    50: [  # QA-003: after row 50 (already modified by V-011)
        {'B': '', 'C': '', 'D': '', 'E': '관리자 A, B 동시에 동일 알림 확인', 'F': '', 'G': '', 'H': 'P2', 'I': '정상 처리됨\n- 에러 없이 양쪽 모두 해당 알림 제거됨', 'O': '리뷰 반영: QA-003 동시조작'},
    ],
    55: [  # QA-009: after row 55
        {'B': '', 'C': '', 'D': '', 'E': '퀵 버튼 빠른 연속 선택', 'F': '', 'G': '', 'H': 'P2', 'I': '마지막 선택한 기간 기준으로 데이터 표시됨\n- 중복 요청/오류 없음', 'O': '리뷰 반영: QA-009 연속입력'},
    ],
    69: [  # QA-004: after row 69
        {'B': '', 'C': '', 'D': '', 'E': '달력 모달 열린 상태에서 브라우저 뒤로가기', 'F': '', 'G': '', 'H': 'P2', 'I': '모달 닫힘\n- 이전 기간 유지됨', 'O': '리뷰 반영: QA-004 뒤로가기'},
    ],
    82: [  # QA-002: after row 82
        {'B': '', 'C': '', 'D': '', 'E': '', 'F': '활용률 0%', 'G': '', 'H': 'P2', 'I': '0% 표시됨\n- 변화율 뱃지 정상 노출됨', 'O': '리뷰 반영: QA-002 경계값'},
        {'B': '', 'C': '', 'D': '', 'E': '', 'F': '활용률 100%', 'G': '', 'H': 'P2', 'I': '100% 표시됨\n- 변화율 뱃지 정상 노출됨', 'O': '리뷰 반영: QA-002 경계값'},
    ],
    123: [  # QA-007: after row 123
        {'B': '', 'C': '', 'D': '', 'E': '', 'F': '수정 완료 후 대시보드 복귀', 'G': '', 'H': 'P2', 'I': '대시보드 데이터 갱신 반영됨\n- 알림 목록/수치 최신 상태로 표시됨', 'O': '리뷰 반영: QA-007 데이터갱신'},
    ],
    202: [  # QA-010: after row 202
        {'B': '', 'C': '', 'D': '세션 만료', 'E': '세션 만료 상태에서 동작 수행', 'F': '', 'G': '', 'H': 'P2', 'I': '로그인 화면으로 리다이렉트됨', 'O': '리뷰 반영: QA-010 세션만료'},
    ],
}

col_map = {'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8, 'I': 9, 'O': 15}

# Step 4: Build final row list
final_rows = []
for r in src_rows:
    rn = r['row_num']
    if rn in delete_rows:
        continue
    final_rows.append(r)
    if rn in new_tcs:
        for tc in new_tcs[rn]:
            new_row = {'data': {i: '' for i in range(1, 16)}, 'styles': None, 'row_num': -1}
            for key, col_idx in col_map.items():
                new_row['data'][col_idx] = tc.get(key, '')
            final_rows.append(new_row)

# Step 5: Write to new sheet
# Column widths
col_widths = {1: 5, 2: 20, 3: 18, 4: 18, 5: 28, 6: 20, 7: 15, 8: 10, 9: 65, 10: 8, 11: 10, 12: 10, 13: 10, 14: 10, 15: 22}
for col, w in col_widths.items():
    ws.column_dimensions[chr(64 + col)].width = w

normal_font = Font(size=10)
header_fill = PatternFill(start_color='FFBFBFBF', end_color='FFBFBFBF', fill_type='solid')
review_fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')  # light yellow for review additions

for idx, r in enumerate(final_rows):
    out_row = idx + 1
    for col in range(1, 16):
        cell = ws.cell(row=out_row, column=col, value=r['data'].get(col, ''))

        if r['styles'] and col in r['styles']:
            src_cell = r['styles'][col]
            copy_cell_style(src_cell, cell)
        else:
            cell.font = normal_font
            cell.border = thin_border
            if col == 8:
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align

        # Highlight review additions in yellow
        if r['row_num'] == -1:
            cell.fill = review_fill

# Copy merged cells for header rows (rows 4-5)
for mc in ws_src.merged_cells.ranges:
    try:
        ws.merge_cells(str(mc))
    except:
        pass

# Update title
ws.cell(row=2, column=2).value = '[centurionsay] Admin v1.3.0 AI 가이드 대시보드 TestCase (리뷰 반영본)'

# Save
out_path = r'c:\Users\kjeon\QA\qa_agent\data\tc\Admin-AI가이드대시보드_리뷰반영.xlsx'
wb.save(out_path)

# Summary
added = sum(1 for r in final_rows if r['row_num'] == -1)
deleted = len(delete_rows)
original = 197
print(f'완료: {out_path}')
print(f'원본 TC: {original}건')
print(f'삭제: {deleted}건 (V-001 중복 Row 118)')
print(f'추가: {added}건 (QA-001~010 리뷰 반영)')
print(f'최종 TC: {original - deleted + added}건')
print(f'리뷰 반영 행은 노란색으로 하이라이트됨')
